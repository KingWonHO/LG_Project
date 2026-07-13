"""engineer_rules 모듈: Engineer_Analysis_Rule.xlsx 로더 (ENG-004 보조).

엔지니어가 규격(RULE_INPUT 시트)으로 xlsx를 업데이트하면 mtime 캐시로 변경 시 자동
재파싱하여 항상 최신 룰을 참고한다. 룰 내용은 자유서술이라 판정엔진이 아니라 LLM
프롬프트 지식으로 사용한다.

[2024 구조 변경] 시트명 RULE_EDITOR → RULE_INPUT, 헤더/값 한글화, enabled/status 삭제.
  헤더: 연번 | 룰 네임 | 데이터 타입 | 룰 스코프 | 해석기준 | 현상정의 | 예상원인 | 필요조치 | Action Type | RAG 키워드
  데이터 타입: 공통/차압/평압,  룰 스코프: 트립/노멀/둘다  (enabled/status 없음 → 모두 사용)
"""

from __future__ import annotations

from pathlib import Path
from threading import Lock

_RULES_PATH = Path(__file__).parent.parent / "data" / "agentData" / "Engineer_Analysis_Rule.xlsx"

# 시트 후보 (신규 우선, 구버전 폴백)
_SHEET_CANDIDATES = ("RULE_INPUT", "RULE_EDITOR")

# 헤더 행 판별: 아래 마커 그룹 중 하나가 한 행에 모두 있으면 룰 테이블 헤더
_HEADER_MARKER_GROUPS = [
    ("룰 네임", "데이터 타입", "룰 스코프"),          # 신규(한글)
    ("rule_name_ko", "data_type", "rule_scope"),        # 구버전(영문) 폴백
]

# 표준 필드 -> 헤더 별칭(부분일치, 소문자 비교). 신규 한글 + 구버전 영문 모두 인식.
_FIELD_ALIASES: dict[str, list[str]] = {
    "rule_no": ["연번"],
    "rule_name_ko": ["룰 네임", "룰네임", "rule_name_ko"],
    "data_type": ["데이터 타입", "데이터타입", "data_type"],
    "rule_scope": ["룰 스코프", "룰스코프", "rule_scope"],
    "analysis_criteria_ko": ["해석기준", "analysis_criteria", "analysis_freeform"],
    "phenomenon_definition_ko": ["현상정의", "phenomenon"],
    "expected_cause_ko": ["예상원인", "expected_cause"],
    "required_action_ko": ["필요조치", "required_action"],
    "action_types": ["action type", "action_type"],
    "rag_keywords": ["rag 키워드", "rag키워드", "rag_keywords"],
}

# 값 매핑 (한글 → 표준 코드)
_DT_MAP = {"공통": "COMMON", "차압": "DPS", "평압": "NODPS"}
_SC_MAP = {"트립": "TRIP", "노멀": "NORMAL", "둘다": "BOTH"}

_cache: list[dict] = []
_cache_mtime: float = -1.0
_lock = Lock()


def _norm_dt(v: object) -> str:
    k = str(v or "").strip()
    if k in _DT_MAP:
        return _DT_MAP[k]
    ku = k.upper()
    return ku if ku in ("COMMON", "DPS", "NODPS") else ""


def _norm_scope(v: object) -> str:
    k = str(v or "").strip()
    if k in _SC_MAP:
        return _SC_MAP[k]
    ku = k.upper()
    return ku if ku in ("TRIP", "NORMAL", "BOTH") else ""


def _parse_rules() -> list[dict]:
    """RULE_INPUT(구: RULE_EDITOR) 시트에서 룰 헤더를 찾아 룰 행들을 dict 리스트로 파싱한다."""
    import openpyxl

    wb = openpyxl.load_workbook(_RULES_PATH, read_only=True, data_only=True)
    sheet = next((s for s in _SHEET_CANDIDATES if s in wb.sheetnames), None)
    if sheet is None:
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))

    # 1) 헤더 행 찾기: 마커 그룹 중 하나가 모두 있는 행
    header_row_idx = -1
    header_cells: dict[int, str] = {}
    for i, r in enumerate(rows):
        texts = {j: str(c).strip() for j, c in enumerate(r) if c is not None and str(c).strip()}
        vals = list(texts.values())
        for group in _HEADER_MARKER_GROUPS:
            if all(any(m in v for v in vals) for m in group):
                header_row_idx, header_cells = i, texts
                break
        if header_row_idx >= 0:
            break
    if header_row_idx < 0:
        return []

    # 2) 표준 필드 -> 컬럼 인덱스 매핑 (부분일치, 소문자)
    field_col: dict[str, int] = {}
    for field, aliases in _FIELD_ALIASES.items():
        for col, text in header_cells.items():
            low = text.lower()
            if any(a in low for a in aliases):
                field_col[field] = col
                break

    name_col = field_col.get("rule_name_ko")
    if name_col is None:
        return []

    # 3) 헤더 다음 행부터 파싱 (룰 네임 비면 skip)
    rules: list[dict] = []
    for r in rows[header_row_idx + 1:]:
        name = r[name_col] if name_col < len(r) else None
        if name is None or str(name).strip() == "":
            continue
        rule: dict = {}
        for field, col in field_col.items():
            val = r[col] if col < len(r) else None
            rule[field] = str(val).strip() if val is not None and str(val).strip() else None
        rules.append(rule)
    return rules


def load_rules(force: bool = False) -> list[dict]:
    """mtime 캐시. 파일이 바뀌면 자동 재파싱, 실패(편집 중 잠금 등) 시 직전 캐시 유지."""
    global _cache, _cache_mtime
    with _lock:
        try:
            mtime = _RULES_PATH.stat().st_mtime
        except OSError:
            return _cache
        if force or mtime != _cache_mtime:
            try:
                _cache = _parse_rules()
                _cache_mtime = mtime
            except Exception:
                pass
        return _cache


def get_applicable_rules(data_type: str | None = None, rule_scope: str | None = None) -> list[dict]:
    """data_type/rule_scope 조건에 맞는 룰만 반환 (신규 구조엔 enabled/status 없음 → 모두 대상).

    data_type: "DPS"/"NODPS" (공통/COMMON 룰은 항상 포함). None이면 필터 없음.
    rule_scope: "TRIP"/"NORMAL" (둘다/BOTH 룰은 항상 포함). None이면 필터 없음.
    """
    out: list[dict] = []
    for r in load_rules():
        rdt = _norm_dt(r.get("data_type"))
        if data_type and rdt not in ("", "COMMON", data_type.upper()):
            continue
        rsc = _norm_scope(r.get("rule_scope"))
        if rule_scope and rsc not in ("", "BOTH", rule_scope.upper()):
            continue
        out.append(r)
    return out


def build_rules_context(data_type: str | None = None, rule_scope: str | None = None) -> str:
    """적용 룰을 LLM 프롬프트용 텍스트로 구성한다 (해석기준/현상정의/예상원인/필요조치)."""
    rules = get_applicable_rules(data_type, rule_scope)
    if not rules:
        return ""
    lines = ["[엔지니어 분석 규칙 (RULE_INPUT, 적용대상)]"]
    for r in rules:
        head = f"■ {r.get('rule_name_ko', '')} ({r.get('data_type', '')}/{r.get('rule_scope', '')})"
        detail = []
        if r.get("analysis_criteria_ko"):
            detail.append(f"  - 해석기준: {r['analysis_criteria_ko']}")
        if r.get("phenomenon_definition_ko"):
            detail.append(f"  - 현상정의: {r['phenomenon_definition_ko']}")
        if r.get("expected_cause_ko"):
            detail.append(f"  - 예상원인(해석): {r['expected_cause_ko']}")
        if r.get("required_action_ko"):
            detail.append(f"  - 필요조치: {r['required_action_ko']}")
        lines.append(head + ("\n" + "\n".join(detail) if detail else ""))
    return "\n".join(lines)
